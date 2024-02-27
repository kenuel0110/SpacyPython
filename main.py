#python_jwt
#gcloud
#sseclient
#Crypto
#pycryptodome
#openpyxl
#xlrd
#progressbar
#requests_toolbelt

#bs4

from bs4 import BeautifulSoup
from urllib.parse import urlparse, urljoin
from datetime import datetime
import requests
from firebase import Firebase
import time
import xlrd
import json
import zipfile
import uuid
import os
import openpyxl
import shutil
import re
import calendar
import progressbar
import sys


#Подключение к Firebase
config = {
  "apiKey": "AIzaSyCZIs6anQ34hAchW0QKwy2LeAmfTBUs-N4",
  "authDomain": "ijournal-fb041.firebaseapp.com",
  "databaseURL": "https://ijournal-fb041-default-rtdb.europe-west1.firebasedatabase.app",
  "projectId": "ijournal-fb041",
  "storageBucket": "ijournal-fb041.appspot.com",
  "messagingSenderId": "673196697696",
  "appId": "1:673196697696:web:e734433541080a1c46fb05",
  "measurementId": "G-2GJJMQLZZB"
}

firebase = Firebase(config)

#storage = firebase.storage()
database = firebase.database()

bar = progressbar.ProgressBar(maxval=8, \
    widgets=[progressbar.Bar('=', '[', ']'), ' ', progressbar.Percentage()])


global number_cycle
number_cycle= ""

listLessons = []

class Lesson:
    def __init__(self, name, type, teacher, dayOfWeek):
        self.name = name 
        self.type = type
        self.teacher = teacher
        self.dayOfWeek = dayOfWeek

listTeacher=[
    "Е.В. Воробьёва",
    "Т.А. Асаева",
    "Ю.И. Арабчикова",
    "Г.Ю. Судакова",
    "Н.В. Гречушкина",
    "И.И. Кулаков",
    "Е.Н. Костылёва",
    "А.А. Бакулина",
    "Т.В. Гончарова",
    "О.В. Тихонова",
]


#Получение данных расписания

chislitelList = []
znamenatelList = []
date_now = datetime.now()

def dateLessons():
    files_list = os.listdir("static/zip/ОЧНИКИ/Расписание занятий")
    excel_file_List = openpyxl.load_workbook(f"static/zip/ОЧНИКИ/Расписание занятий/{files_list[0]}")
    sheetNames = excel_file_List.sheetnames
    listInf_sheet = excel_file_List[sheetNames[3]]
    dateString = listInf_sheet["C67"].value         # Получение данных (знаменатель/числитель) из таблицы
    dateString = dateString.split()
    dateString = dateString[8:]

    cur = 0
    while cur < len(dateString):
        if (cur % 2 == 0):
            chislitelList.append(dateString[cur])
        else:
            znamenatelList.append(dateString[cur])
        cur += 1
    
    getCycleNum(excel_file_List)

def getCycleNum(excel_file_List):
    date_now_ = datetime.now().strftime('%d/%m/%Y')
    year = date_now_.split("/")[2]
    date_now_ = datetime.strptime(date_now_, '%d/%m/%Y')
    key = "none"
    for i in chislitelList:
        item = i.split("-")
        convert1 = f"{item[0].replace('.', '/')}/{year}"
        convert2 = f"{item[1].replace('.', '/')}/{year}"
        if(datetime.strptime(convert1, "%d/%m/%Y") <= date_now_ <= datetime.strptime(convert2, "%d/%m/%Y")):
            key = "chis"
            continue

        if (key == "none" or key == "chis"):
            for i in znamenatelList:
                item = i.split("-")
                convert1 = f"{item[0].replace('.', '/')}/{year}"
                convert2 = f"{item[1].replace('.', '/')}/{year}"
            if(datetime.strptime(convert1, "%d/%m/%Y") <= date_now <= datetime.strptime(convert2, "%d/%m/%Y")):
                    key = "znamen"
                    continue
    bar.update(5)
    getDateZnOrCh(excel_file_List)


def getDateZnOrCh(excel_file_List):
    global number_cycle
    date_now = datetime.now()
    year = date_now.strftime("%d.%m.%y").split(".")[2]
    date_now = date_now.strptime(date_now.strftime("%d/%m/%y"), "%d/%m/%y")
    for i in chislitelList:
        item = i.split('-')
        convert1 = f"{item[0].replace('.', '/')}/{year}"
        convert2 = f"{item[1].replace('.', '/')}/{year}"
        if(datetime.strptime(convert1, "%d/%m/%y") <= date_now <= datetime.strptime(convert2, "%d/%m/%y")):
            number_cycle = "chis"
            break
    
    if (number_cycle != "chis"):
            for i in znamenatelList:
                item = i.split('-')
                convert1 = f"{item[0].replace('.', '/')}/{year}"
                convert2 = f"{item[1].replace('.', '/')}/{year}"
                if(datetime.strptime(convert1, "%d/%m/%y") <= date_now <= datetime.strptime(convert2, "%d/%m/%y")):
                    number_cycle = "znamen"
                    break
    bar.update(6)
    getLessons(excel_file_List)


def getLessons(excel_file_List):
    global number_cycle
    listLessons.clear()
    sheetNames = excel_file_List.sheetnames
    listInf_sheet = excel_file_List[sheetNames[3]]

    cursor = 0
    bar.update(7)
    if(number_cycle == "chis"):
        cursor = 0
        while cursor < 8:
            item_mon = str(listInf_sheet[f"C{11 + cursor}"].value)
            item_mon_list = re.sub(r'\s+', ' ', item_mon)
            item_mon_list = item_mon_list.split(' ')

            item_tue = str(listInf_sheet[f"C{20 + cursor}"].value)
            item_tue_list = re.sub(r'\s+', ' ', item_tue)
            item_tue_list = item_tue_list.split(' ')

            item_wed = str(listInf_sheet[f"C{29 + cursor}"].value)
            item_wed_list = re.sub(r'\s+', ' ', item_wed)
            item_wed_list = item_wed_list.split(' ')

            item_thu = str(listInf_sheet[f"C{38 + cursor}"].value)
            item_thu_list = re.sub(r'\s+', ' ', item_thu)
            item_thu_list = item_thu_list.split(' ')

            item_fri = str(listInf_sheet[f"C{47 + cursor}"].value)
            item_fri_list = re.sub(r'\s+', ' ', item_fri)
            item_fri_list = item_fri_list.split(' ')

            item_sat = str(listInf_sheet[f"C{56 + cursor}"].value)
            item_sat_list = re.sub(r'\s+', ' ', item_sat)
            item_sat_list = item_sat_list.split(' ')

            if (item_mon != "None"):
                if (item_mon.find("Иностранный язык") != -1):
                    listLessons.append(Lesson(f"{item_mon_list[0]} {item_mon_list[1]}", item_mon_list[2], item_mon_list[3], "Mon"))
                elif (item_mon.find("Экология") != -1):
                    listLessons.append(Lesson(f"{item_mon_list[0]}", item_mon_list[1], item_mon_list[2], "Mon"))
                elif (item_mon.find("Математика") != -1):
                    listLessons.append(Lesson(f"{item_mon_list[0]}", item_mon_list[1], item_mon_list[2], "Mon"))
                elif (item_mon.find("Интернет-технологии") != -1):
                    listLessons.append(Lesson(f"{item_mon_list[0]}", item_mon_list[1], item_mon_list[2], "Mon"))
                elif (item_mon.find("Введение в профессию") != -1):
                    listLessons.append(Lesson(f"{item_mon_list[0]} {item_mon_list[1]} {item_mon_list[2]}", item_mon_list[3], item_mon_list[4], "Mon"))
                elif (item_mon.find("Физическая культура и спорт") != -1):
                    listLessons.append(Lesson(f"{item_mon_list[0]} {item_mon_list[1]} {item_mon_list[2]} {item_mon_list[3]}", item_mon_list[4], item_mon_list[5], "Mon"))
                elif (item_mon.find("История (история России, всеобщая история)") != -1):
                    listLessons.append(Lesson(f"{item_mon_list[0]}", item_mon_list[5], item_mon_list[6], "Mon"))
                elif (item_mon.find("Введение в проектную деятельность") != -1):
                    listLessons.append(Lesson(f"{item_mon_list[0]} {item_mon_list[1]} {item_mon_list[2]} {item_mon_list[3]}", item_mon_list[4], item_mon_list[5], "Mon"))
                elif (item_mon.find("Введение в информационные технологии") != -1):
                    listLessons.append(Lesson(f"{item_mon_list[0]} {item_mon_list[1]} {item_mon_list[2]} {item_mon_list[3]}", item_mon_list[4], item_mon_list[5], "Mon"))
                elif (item_mon.find("Русский язык и культура речи") != -1):
                    listLessons.append(Lesson(f"{item_mon_list[0]} {item_mon_list[1]}", item_mon_list[5], item_mon_list[6], "Mon"))
                else:
                    pass
            else:
                pass

            if (item_tue != "None"):
                if (item_tue.find("Иностранный язык") != -1):
                    listLessons.append(Lesson(f"{item_tue_list[0]} {item_tue_list[1]}", item_tue_list[2], item_tue_list[3], "Tue"))
                elif (item_tue.find("Экология") != -1):
                    listLessons.append(Lesson(f"{item_tue_list[0]}", item_tue_list[1], item_tue_list[2], "Tue"))
                elif (item_tue.find("Математика") != -1):
                    listLessons.append(Lesson(f"{item_tue_list[0]}", item_tue_list[1], item_tue_list[2], "Tue"))
                elif (item_tue.find("Интернет-технологии") != -1):
                    listLessons.append(Lesson(f"{item_tue_list[0]}", item_tue_list[1], item_tue_list[2], "Tue"))
                elif (item_tue.find("Введение в профессию") != -1):
                    listLessons.append(Lesson(f"{item_tue_list[0]} {item_tue_list[1]} {item_tue_list[2]}", item_tue_list[3], item_tue_list[4], "Tue"))
                elif (item_tue.find("Физическая культура и спорт") != -1):
                    listLessons.append(Lesson(f"{item_tue_list[0]} {item_tue_list[1]} {item_tue_list[2]} {item_tue_list[3]}", item_tue_list[4], item_tue_list[5], "Tue"))
                elif (item_tue.find("История (история России, всеобщая история)") != -1):
                    listLessons.append(Lesson(f"{item_tue_list[0]}", item_tue_list[5], item_tue_list[6], "Tue"))
                elif (item_tue.find("Введение в проектную деятельность") != -1):
                    listLessons.append(Lesson(f"{item_tue_list[0]} {item_tue_list[1]} {item_tue_list[2]} {item_tue_list[3]}", item_tue_list[4], item_tue_list[5], "Tue"))
                elif (item_tue.find("Введение в информационные технологии") != -1):
                    listLessons.append(Lesson(f"{item_tue_list[0]} {item_tue_list[1]} {item_tue_list[2]} {item_tue_list[3]}", item_tue_list[4], item_tue_list[5], "Tue"))
                elif (item_tue.find("Русский язык и культура речи") != -1):
                    listLessons.append(Lesson(f"{item_tue_list[0]} {item_tue_list[1]}", item_tue_list[5], item_tue_list[6], "Tue"))
                else:
                    pass
            else:
                pass

            if (item_wed != "None"):
                if (item_wed.find("Иностранный язык") != -1):
                    listLessons.append(Lesson(f"{item_wed_list[0]} {item_wed_list[1]}", item_wed_list[2], item_wed_list[3], "Wed"))
                elif (item_wed.find("Экология") != -1):
                    listLessons.append(Lesson(f"{item_wed_list[0]}", item_wed_list[1], item_wed_list[2], "Wed"))
                elif (item_wed.find("Математика") != -1):
                    listLessons.append(Lesson(f"{item_wed_list[0]}", item_wed_list[1], item_wed_list[2], "Wed"))
                elif (item_wed.find("Интернет-технологии") != -1):
                    listLessons.append(Lesson(f"{item_wed_list[0]}", item_wed_list[1], item_wed_list[2], "Wed"))
                elif (item_wed.find("Введение в профессию") != -1):
                    listLessons.append(Lesson(f"{item_wed_list[0]} {item_wed_list[1]} {item_wed_list[2]}", item_wed_list[3], item_wed_list[4], "Wed"))
                elif (item_wed.find("Физическая культура и спорт") != -1):
                    listLessons.append(Lesson(f"{item_wed_list[0]} {item_wed_list[1]} {item_wed_list[2]} {item_wed_list[3]}", item_wed_list[4], item_wed_list[5], "Wed"))
                elif (item_wed.find("История (история России, всеобщая история)") != -1):
                    listLessons.append(Lesson(f"{item_wed_list[0]}", item_wed_list[5], item_wed_list[6], "Wed"))
                elif (item_wed.find("Введение в проектную деятельность") != -1):
                    listLessons.append(Lesson(f"{item_wed_list[0]} {item_wed_list[1]} {item_wed_list[2]} {item_wed_list[3]}", item_wed_list[4], item_wed_list[5], "Wed"))
                elif (item_wed.find("Введение в информационные технологии") != -1):
                    listLessons.append(Lesson(f"{item_wed_list[0]} {item_wed_list[1]} {item_wed_list[2]} {item_wed_list[3]}", item_wed_list[4], item_wed_list[5], "Wed"))
                elif (item_wed.find("Русский язык и культура речи") != -1):
                    listLessons.append(Lesson(f"{item_wed_list[0]} {item_wed_list[1]}", item_wed_list[5], item_wed_list[6], "Wed"))
                else:
                    pass
            else:
                pass

            if (item_thu != "None"):
                if (item_thu.find("Иностранный язык") != -1):
                    listLessons.append(Lesson(f"{item_thu_list[0]} {item_thu_list[1]}", item_thu_list[2], item_thu_list[3], "Thu"))
                elif (item_thu.find("Экология") != -1):
                    listLessons.append(Lesson(f"{item_thu_list[0]}", item_thu_list[1], item_thu_list[2], "Thu"))
                elif (item_thu.find("Математика") != -1):
                    listLessons.append(Lesson(f"{item_thu_list[0]}", item_thu_list[1], item_thu_list[2], "Thu"))
                elif (item_thu.find("Интернет-технологии") != -1):
                    listLessons.append(Lesson(f"{item_thu_list[0]}", item_thu_list[1], item_thu_list[2], "Thu"))
                elif (item_thu.find("Введение в профессию") != -1):
                    listLessons.append(Lesson(f"{item_thu_list[0]} {item_thu_list[1]} {item_thu_list[2]}", item_thu_list[3], item_thu_list[4], "Thu"))
                elif (item_thu.find("Физическая культура и спорт") != -1):
                    listLessons.append(Lesson(f"{item_thu_list[0]} {item_thu_list[1]} {item_thu_list[2]} {item_thu_list[3]}", item_thu_list[4], item_thu_list[5], "Thu"))
                elif (item_thu.find("История (история России, всеобщая история)") != -1):
                    listLessons.append(Lesson(f"{item_thu_list[0]}", item_thu_list[5], item_thu_list[6], "Thu"))
                elif (item_thu.find("Введение в проектную деятельность") != -1):
                    listLessons.append(Lesson(f"{item_thu_list[0]} {item_thu_list[1]} {item_thu_list[2]} {item_thu_list[3]}", item_thu_list[4], item_thu_list[5], "Thu"))
                elif (item_thu.find("Введение в информационные технологии") != -1):
                    listLessons.append(Lesson(f"{item_thu_list[0]} {item_thu_list[1]} {item_thu_list[2]} {item_thu_list[3]}", item_thu_list[4], item_thu_list[5], "Thu"))
                elif (item_thu.find("Русский язык и культура речи") != -1):
                    listLessons.append(Lesson(f"{item_thu_list[0]} {item_thu_list[1]}", item_thu_list[5], item_thu_list[6], "Thu"))
                else:
                    pass
            else:
                pass

            if (item_fri != "None"):
                if (item_fri.find("Иностранный язык") != -1):
                    listLessons.append(Lesson(f"{item_fri_list[0]} {item_fri_list[1]}", item_fri_list[2], item_fri_list[3], "Fri"))
                elif (item_fri.find("Экология") != -1):
                    listLessons.append(Lesson(f"{item_fri_list[0]}", item_fri_list[1], item_fri_list[2], "Fri"))
                elif (item_fri.find("Математика") != -1):
                    listLessons.append(Lesson(f"{item_fri_list[0]}", item_fri_list[1], item_fri_list[2], "Fri"))
                elif (item_fri.find("Интернет-технологии") != -1):
                    listLessons.append(Lesson(f"{item_fri_list[0]}", item_fri_list[1], item_fri_list[2], "Fri"))
                elif (item_fri.find("Введение в профессию") != -1):
                    listLessons.append(Lesson(f"{item_fri_list[0]} {item_fri_list[1]} {item_fri_list[2]}", item_fri_list[3], item_fri_list[4], "Fri"))
                elif (item_fri.find("Физическая культура и спорт") != -1):
                    listLessons.append(Lesson(f"{item_fri_list[0]} {item_fri_list[1]} {item_fri_list[2]} {item_fri_list[3]}", item_fri_list[4], item_fri_list[5], "Fri"))
                elif (item_fri.find("История (история России, всеобщая история)") != -1):
                    listLessons.append(Lesson(f"{item_fri_list[0]}", item_fri_list[5], item_fri_list[6], "Fri"))
                elif (item_fri.find("Введение в проектную деятельность") != -1):
                    listLessons.append(Lesson(f"{item_fri_list[0]} {item_fri_list[1]} {item_fri_list[2]} {item_fri_list[3]}", item_fri_list[4], item_fri_list[5], "Fri"))
                elif (item_fri.find("Введение в информационные технологии") != -1):
                    listLessons.append(Lesson(f"{item_fri_list[0]} {item_fri_list[1]} {item_fri_list[2]} {item_fri_list[3]}", item_fri_list[4], item_fri_list[5], "Fri"))
                elif (item_fri.find("Русский язык и культура речи") != -1):
                    listLessons.append(Lesson(f"{item_fri_list[0]} {item_fri_list[1]}", item_fri_list[5], item_fri_list[6], "Fri"))
                else:
                    pass
            else:
                pass

            if (item_sat != "None"):
                if (item_sat.find("Иностранный язык") != -1):
                    listLessons.append(Lesson(f"{item_sat_list[0]} {item_sat_list[1]}", item_sat_list[2], item_sat_list[3], "Sat"))
                elif (item_sat.find("Экология") != -1):
                    listLessons.append(Lesson(f"{item_sat_list[0]}", item_sat_list[1], item_sat_list[2], "Sat"))
                elif (item_sat.find("Математика") != -1):
                    listLessons.append(Lesson(f"{item_sat_list[0]}", item_sat_list[1], item_sat_list[2], "Sat"))
                elif (item_sat.find("Интернет-технологии") != -1):
                    listLessons.append(Lesson(f"{item_sat_list[0]}", item_sat_list[1], item_sat_list[2], "Sat"))
                elif (item_sat.find("Введение в профессию") != -1):
                    listLessons.append(Lesson(f"{item_sat_list[0]} {item_sat_list[1]} {item_sat_list[2]}", item_sat_list[3], item_sat_list[4], "Sat"))
                elif (item_sat.find("Физическая культура и спорт") != -1):
                    listLessons.append(Lesson(f"{item_sat_list[0]} {item_sat_list[1]} {item_sat_list[2]} {item_sat_list[3]}", item_sat_list[4], item_sat_list[5], "Sat"))
                elif (item_sat.find("История (история России, всеобщая история)") != -1):
                    listLessons.append(Lesson(f"{item_sat_list[0]}", item_sat_list[5], item_sat_list[6], "Sat"))
                elif (item_sat.find("Введение в проектную деятельность") != -1):
                    listLessons.append(Lesson(f"{item_sat_list[0]} {item_sat_list[1]} {item_sat_list[2]} {item_sat_list[3]}", item_sat_list[4], item_sat_list[5], "Sat"))
                elif (item_sat.find("Введение в информационные технологии") != -1):
                    listLessons.append(Lesson(f"{item_sat_list[0]} {item_sat_list[1]} {item_sat_list[2]} {item_sat_list[3]}", item_sat_list[4], item_sat_list[5], "Sat"))
                elif (item_sat.find("Русский язык и культура речи") != -1):
                    listLessons.append(Lesson(f"{item_sat_list[0]} {item_sat_list[1]}", item_sat_list[5], item_sat_list[6], "Sat"))
                else:
                    pass
            else:
                pass

            cursor += 2

    elif(number_cycle == "znamen"):
        cursor = -1

        while cursor < 8:
            item_mon = str(listInf_sheet[f"C{11 + cursor}"].value)
            item_mon_list = re.sub(r'\s+', ' ', item_mon)
            item_mon_list = item_mon.split(' ')

            item_tue = str(listInf_sheet[f"C{20 + cursor}"].value)
            item_tue_list = re.sub(r'\s+', ' ', item_tue)
            item_tue_list = item_tue_list.split(' ')

            item_wed = str(listInf_sheet[f"C{29 + cursor}"].value)
            item_wed_list = re.sub(r'\s+', ' ', item_wed)
            item_wed_list = item_wed_list.split(' ')

            item_thu = str(listInf_sheet[f"C{38 + cursor}"].value)
            item_thu_list = re.sub(r'\s+', ' ', item_thu)
            item_thu_list = item_thu_list.split(' ')

            item_fri = str(listInf_sheet[f"C{47 + cursor}"].value)
            item_fri_list = re.sub(r'\s+', ' ', item_fri)
            item_fri_list = item_fri_list.split(' ')

            item_sat = str(listInf_sheet[f"C{56 + cursor}"].value)
            item_sat_list = re.sub(r'\s+', ' ', item_sat)
            item_sat_list = item_sat_list.split(' ')

            if (item_mon != "None"):
                if (item_mon.find("Иностранный язык") != -1):
                    listLessons.append(Lesson(f"{item_mon_list[0]} {item_mon_list[1]}", item_mon_list[2], item_mon_list[3], "Mon"))
                elif (item_mon.find("Экология") != -1):
                    listLessons.append(Lesson(f"{item_mon_list[0]}", item_mon_list[1], item_mon_list[2], "Mon"))
                elif (item_mon.find("Математика") != -1):
                    listLessons.append(Lesson(f"{item_mon_list[0]}", item_mon_list[1], item_mon_list[2], "Mon"))
                elif (item_mon.find("Интернет-технологии") != -1):
                    listLessons.append(Lesson(f"{item_mon_list[0]}", item_mon_list[1], item_mon_list[2], "Mon"))
                elif (item_mon.find("Введение в профессию") != -1):
                    listLessons.append(Lesson(f"{item_mon_list[0]} {item_mon_list[1]} {item_mon_list[2]}", item_mon_list[3], item_mon_list[4], "Mon"))
                elif (item_mon.find("Физическая культура и спорт") != -1):
                    listLessons.append(Lesson(f"{item_mon_list[0]} {item_mon_list[1]} {item_mon_list[2]} {item_mon_list[3]}", item_mon_list[4], item_mon_list[5], "Mon"))
                elif (item_mon.find("История (история России, всеобщая история)") != -1):
                    listLessons.append(Lesson(f"{item_mon_list[0]}", item_mon_list[5], item_mon_list[6], "Mon"))
                elif (item_mon.find("Введение в проектную деятельность") != -1):
                    listLessons.append(Lesson(f"{item_mon_list[0]} {item_mon_list[1]} {item_mon_list[2]} {item_mon_list[3]}", item_mon_list[4], item_mon_list[5], "Mon"))
                elif (item_mon.find("Введение в информационные технологии") != -1):
                    listLessons.append(Lesson(f"{item_mon_list[0]} {item_mon_list[1]} {item_mon_list[2]} {item_mon_list[3]}", item_mon_list[4], item_mon_list[5], "Mon"))
                elif (item_mon.find("Русский язык и культура речи") != -1):
                    listLessons.append(Lesson(f"{item_mon_list[0]} {item_mon_list[1]}", item_mon_list[5], item_mon_list[6], "Mon"))
                else:
                    pass
            else:
                pass

            if (item_tue != "None"):
                if (item_tue.find("Иностранный язык") != -1):
                    listLessons.append(Lesson(f"{item_tue_list[0]} {item_tue_list[1]}", item_tue_list[2], item_tue_list[3], "Tue"))
                elif (item_tue.find("Экология") != -1):
                    listLessons.append(Lesson(f"{item_tue_list[0]}", item_tue_list[1], item_tue_list[2], "Tue"))
                elif (item_tue.find("Математика") != -1):
                    listLessons.append(Lesson(f"{item_tue_list[0]}", item_tue_list[1], item_tue_list[2], "Tue"))
                elif (item_tue.find("Интернет-технологии") != -1):
                    listLessons.append(Lesson(f"{item_tue_list[0]}", item_tue_list[1], item_tue_list[2], "Tue"))
                elif (item_tue.find("Введение в профессию") != -1):
                    listLessons.append(Lesson(f"{item_tue_list[0]} {item_tue_list[1]} {item_tue_list[2]}", item_tue_list[3], item_tue_list[4], "Tue"))
                elif (item_tue.find("Физическая культура и спорт") != -1):
                    listLessons.append(Lesson(f"{item_tue_list[0]} {item_tue_list[1]} {item_tue_list[2]} {item_tue_list[3]}", item_tue_list[4], item_tue_list[5], "Tue"))
                elif (item_tue.find("История (история России, всеобщая история)") != -1):
                    listLessons.append(Lesson(f"{item_tue_list[0]}", item_tue_list[5], item_tue_list[6], "Tue"))
                elif (item_tue.find("Введение в проектную деятельность") != -1):
                    listLessons.append(Lesson(f"{item_tue_list[0]} {item_tue_list[1]} {item_tue_list[2]} {item_tue_list[3]}", item_tue_list[4], item_tue_list[5], "Tue"))
                elif (item_tue.find("Введение в информационные технологии") != -1):
                    listLessons.append(Lesson(f"{item_tue_list[0]} {item_tue_list[1]} {item_tue_list[2]} {item_tue_list[3]}", item_tue_list[4], item_tue_list[5], "Tue"))
                elif (item_tue.find("Русский язык и культура речи") != -1):
                    listLessons.append(Lesson(f"{item_tue_list[0]} {item_tue_list[1]}", item_tue_list[5], item_tue_list[6], "Tue"))
                else:
                    pass
            else:
                pass

            if (item_wed != "None"):
                if (item_wed.find("Иностранный язык") != -1):
                    listLessons.append(Lesson(f"{item_wed_list[0]} {item_wed_list[1]}", item_wed_list[2], item_wed_list[3], "Wed"))
                elif (item_wed.find("Экология") != -1):
                    listLessons.append(Lesson(f"{item_wed_list[0]}", item_wed_list[1], item_wed_list[2], "Wed"))
                elif (item_wed.find("Математика") != -1):
                    listLessons.append(Lesson(f"{item_wed_list[0]}", item_wed_list[1], item_wed_list[2], "Wed"))
                elif (item_wed.find("Интернет-технологии") != -1):
                    listLessons.append(Lesson(f"{item_wed_list[0]}", item_wed_list[1], item_wed_list[2], "Wed"))
                elif (item_wed.find("Введение в профессию") != -1):
                    listLessons.append(Lesson(f"{item_wed_list[0]} {item_wed_list[1]} {item_wed_list[2]}", item_wed_list[3], item_wed_list[4], "Wed"))
                elif (item_wed.find("Физическая культура и спорт") != -1):
                    listLessons.append(Lesson(f"{item_wed_list[0]} {item_wed_list[1]} {item_wed_list[2]} {item_wed_list[3]}", item_wed_list[4], item_wed_list[5], "Wed"))
                elif (item_wed.find("История (история России, всеобщая история)") != -1):
                    listLessons.append(Lesson(f"{item_wed_list[0]}", item_wed_list[5], item_wed_list[6], "Wed"))
                elif (item_wed.find("Введение в проектную деятельность") != -1):
                    listLessons.append(Lesson(f"{item_wed_list[0]} {item_wed_list[1]} {item_wed_list[2]} {item_wed_list[3]}", item_wed_list[4], item_wed_list[5], "Wed"))
                elif (item_wed.find("Введение в информационные технологии") != -1):
                    listLessons.append(Lesson(f"{item_wed_list[0]} {item_wed_list[1]} {item_wed_list[2]} {item_wed_list[3]}", item_wed_list[4], item_wed_list[5], "Wed"))
                elif (item_wed.find("Русский язык и культура речи") != -1):
                    listLessons.append(Lesson(f"{item_wed_list[0]} {item_wed_list[1]}", item_wed_list[5], item_wed_list[6], "Wed"))
                else:
                    pass
            else:
                pass

            if (item_thu != "None"):
                if (item_thu.find("Иностранный язык") != -1):
                    listLessons.append(Lesson(f"{item_thu_list[0]} {item_thu_list[1]}", item_thu_list[2], item_thu_list[3], "Thu"))
                elif (item_thu.find("Экология") != -1):
                    listLessons.append(Lesson(f"{item_thu_list[0]}", item_thu_list[1], item_thu_list[2], "Thu"))
                elif (item_thu.find("Математика") != -1):
                    listLessons.append(Lesson(f"{item_thu_list[0]}", item_thu_list[1], item_thu_list[2], "Thu"))
                elif (item_thu.find("Интернет-технологии") != -1):
                    listLessons.append(Lesson(f"{item_thu_list[0]}", item_thu_list[1], item_thu_list[2], "Thu"))
                elif (item_thu.find("Введение в профессию") != -1):
                    listLessons.append(Lesson(f"{item_thu_list[0]} {item_thu_list[1]} {item_thu_list[2]}", item_thu_list[3], item_thu_list[4], "Thu"))
                elif (item_thu.find("Физическая культура и спорт") != -1):
                    listLessons.append(Lesson(f"{item_thu_list[0]} {item_thu_list[1]} {item_thu_list[2]} {item_thu_list[3]}", item_thu_list[4], item_thu_list[5], "Thu"))
                elif (item_thu.find("История (история России, всеобщая история)") != -1):
                    listLessons.append(Lesson(f"{item_thu_list[0]}", item_thu_list[5], item_thu_list[6], "Thu"))
                elif (item_thu.find("Введение в проектную деятельность") != -1):
                    listLessons.append(Lesson(f"{item_thu_list[0]} {item_thu_list[1]} {item_thu_list[2]} {item_thu_list[3]}", item_thu_list[4], item_thu_list[5], "Thu"))
                elif (item_thu.find("Введение в информационные технологии") != -1):
                    listLessons.append(Lesson(f"{item_thu_list[0]} {item_thu_list[1]} {item_thu_list[2]} {item_thu_list[3]}", item_thu_list[4], item_thu_list[5], "Thu"))
                elif (item_thu.find("Русский язык и культура речи") != -1):
                    listLessons.append(Lesson(f"{item_thu_list[0]} {item_thu_list[1]}", item_thu_list[5], item_thu_list[6], "Thu"))
                else:
                    pass
            else:
                pass

            if (item_fri != "None"):
                if (item_fri.find("Иностранный язык") != -1):
                    listLessons.append(Lesson(f"{item_fri_list[0]} {item_fri_list[1]}", item_fri_list[2], item_fri_list[3], "Fri"))
                elif (item_fri.find("Экология") != -1):
                    listLessons.append(Lesson(f"{item_fri_list[0]}", item_fri_list[1], item_fri_list[2], "Fri"))
                elif (item_fri.find("Математика") != -1):
                    listLessons.append(Lesson(f"{item_fri_list[0]}", item_fri_list[1], item_fri_list[2], "Fri"))
                elif (item_fri.find("Интернет-технологии") != -1):
                    listLessons.append(Lesson(f"{item_fri_list[0]}", item_fri_list[1], item_fri_list[2], "Fri"))
                elif (item_fri.find("Введение в профессию") != -1):
                    listLessons.append(Lesson(f"{item_fri_list[0]} {item_fri_list[1]} {item_fri_list[2]}", item_fri_list[3], item_fri_list[4], "Fri"))
                elif (item_fri.find("Физическая культура и спорт") != -1):
                    listLessons.append(Lesson(f"{item_fri_list[0]} {item_fri_list[1]} {item_fri_list[2]} {item_fri_list[3]}", item_fri_list[4], item_fri_list[5], "Fri"))
                elif (item_fri.find("История (история России, всеобщая история)") != -1):
                    listLessons.append(Lesson(f"{item_fri_list[0]}", item_fri_list[5], item_fri_list[6], "Fri"))
                elif (item_fri.find("Введение в проектную деятельность") != -1):
                    listLessons.append(Lesson(f"{item_fri_list[0]} {item_fri_list[1]} {item_fri_list[2]} {item_fri_list[3]}", item_fri_list[4], item_fri_list[5], "Fri"))
                elif (item_fri.find("Введение в информационные технологии") != -1):
                    listLessons.append(Lesson(f"{item_fri_list[0]} {item_fri_list[1]} {item_fri_list[2]} {item_fri_list[3]}", item_fri_list[4], item_fri_list[5], "Fri"))
                elif (item_fri.find("Русский язык и культура речи") != -1):
                    listLessons.append(Lesson(f"{item_fri_list[0]} {item_fri_list[1]}", item_fri_list[5], item_fri_list[6], "Fri"))
                else:
                    pass
            else:
                pass

            if (item_sat != "None"):
                if (item_sat.find("Иностранный язык") != -1):
                    listLessons.append(Lesson(f"{item_sat_list[0]} {item_sat_list[1]}", item_sat_list[2], item_sat_list[3], "Sat"))
                elif (item_sat.find("Экология") != -1):
                    listLessons.append(Lesson(f"{item_sat_list[0]}", item_sat_list[1], item_sat_list[2], "Sat"))
                elif (item_sat.find("Математика") != -1):
                    listLessons.append(Lesson(f"{item_sat_list[0]}", item_sat_list[1], item_sat_list[2], "Sat"))
                elif (item_sat.find("Интернет-технологии") != -1):
                    listLessons.append(Lesson(f"{item_sat_list[0]}", item_sat_list[1], item_sat_list[2], "Sat"))
                elif (item_sat.find("Введение в профессию") != -1):
                    listLessons.append(Lesson(f"{item_sat_list[0]} {item_sat_list[1]} {item_sat_list[2]}", item_sat_list[3], item_sat_list[4], "Sat"))
                elif (item_sat.find("Физическая культура и спорт") != -1):
                    listLessons.append(Lesson(f"{item_sat_list[0]} {item_sat_list[1]} {item_sat_list[2]} {item_sat_list[3]}", item_sat_list[4], item_sat_list[5], "Sat"))
                elif (item_sat.find("История (история России, всеобщая история)") != -1):
                    listLessons.append(Lesson(f"{item_sat_list[0]}", item_sat_list[5], item_sat_list[6], "Sat"))
                elif (item_sat.find("Введение в проектную деятельность") != -1):
                    listLessons.append(Lesson(f"{item_sat_list[0]} {item_sat_list[1]} {item_sat_list[2]} {item_sat_list[3]}", item_sat_list[4], item_sat_list[5], "Sat"))
                elif (item_sat.find("Введение в информационные технологии") != -1):
                    listLessons.append(Lesson(f"{item_sat_list[0]} {item_sat_list[1]} {item_sat_list[2]} {item_sat_list[3]}", item_sat_list[4], item_sat_list[5], "Sat"))
                elif (item_sat.find("Русский язык и культура речи") != -1):
                    listLessons.append(Lesson(f"{item_sat_list[0]} {item_sat_list[1]}", item_sat_list[5], item_sat_list[6], "Sat"))
                else:
                    pass
            else:
                pass
            cursor += 2
    database.child("shedule").remove()
    pushLesson2Firebase()

def pushLesson2Firebase():
    for i in listLessons:
            database.child("shedule").push({
                    "name": f"{i.name}",
                    "type": f"{i.type}",
                    "teacher": f"{i.teacher}",
                    "dayOfWeek": f"{i.dayOfWeek}"
                    })
    bar.update(8)
    bar.finish

def getDataTimetable():
    bar.start()
    parsing()
    bar.update(1)
    download_file(internal_urls[2])
    bar.update(2)
    zipOpen()
    bar.update(3)
    dateLessons()
    bar.update(4)






#Скачивание данных
def parsing():
    internal_urls.clear()
    url = "http://own.rimsou.ru/owncloud/index.php/s/VidRmPFrstIhsj9?path=%2FРасписание%20занятий"
    domain_name = urlparse(url).netloc
    soup = BeautifulSoup(requests.get(url).content, "html.parser")
    for a_tag in soup.findAll("a"):
        href = a_tag.attrs.get("href")
        if href == "" or href is None:
            # пустой тег href
            continue
        if href in internal_urls:
            # уже в наборе
            continue
        internal_urls.append(href)



def download_file(link):
    r = requests.get(link)
    if (os.path.isdir('static') != True):
        os.mkdir("static")
    else:
        shutil.rmtree('/static', ignore_errors=True)
    name = 'static/очники.zip'
    file = open(name, 'wb')
    file.write(r.content)
    file.close


def zipOpen():
    archive  = 'static/очники.zip'
    if (os.path.isdir('static/zip') != True):
        os.mkdir('static/zip')
    with zipfile.ZipFile(archive, 'r') as zip_file:
        zip_file.extractall('static/zip')



#Получение данных кабинета
class Teachers_name:
    def __init__(self, name, rowx, colx):
        self.name = name
        self.rowx = rowx 
        self.colx = colx 

class Cabinets:
    def __init__(self, name, rowx_start, colx):
        self.number = name
        self.rowx_start = rowx_start 
        self.colx = colx 
    

cabinets = [
    Cabinets(2, 2, 2),
    Cabinets(3, 9, 2),
    Cabinets(4, 16, 2),
    Cabinets(5, 23, 2),
    Cabinets(12, 30, 2),
    Cabinets(13, 37, 2),
    Cabinets(14, 44, 2),
    Cabinets(15, 51, 2),
    Cabinets(113, 65, 2),
    Cabinets(115, 72, 2),
    Cabinets(116, 79, 2),
    Cabinets(117, 86, 2),
    Cabinets(118, 93, 2),
    Cabinets(205, 100, 2),
    Cabinets(206, 107, 2),
    Cabinets(207, 114, 2),
    Cabinets(208, 121, 2),
    Cabinets(204, 128, 2),
    Cabinets("Акт. Зал", 135, 2),
    Cabinets(16, 2, 5),
    Cabinets(25, 9, 5),
    Cabinets(26, 16, 5),
    Cabinets(27, 23, 5),
    Cabinets(28, 30, 5),
    Cabinets(29, 37, 5),
    Cabinets("29(a)", 44, 5),
    Cabinets(31, 51, 5),
    Cabinets(109, 58, 5),
    Cabinets(209, 65, 5),
    Cabinets(210, 72, 5),
    Cabinets(211, 79, 5),
    Cabinets(212, 86, 5),
    Cabinets(213, 93, 5),
    Cabinets(214, 100, 5),
    Cabinets(216, 107, 5),
    Cabinets(217, 114, 5),
    Cabinets(220, 121, 5),
    Cabinets(221, 128, 5),
    Cabinets(222, 135, 5),


    Cabinets(6, 145, 2),
    Cabinets(7, 152, 2),
    Cabinets(21, 159, 2),
    Cabinets(22, 166, 2),
    Cabinets(23, 173, 2),
    Cabinets(24, 180, 2),
    Cabinets(25, 187, 2),

    Cabinets(2, 145, 5),
    Cabinets(31, 152, 5),
    Cabinets(32, 159, 5),
    Cabinets(33, 166, 5),
    Cabinets(34, 173, 5),
    Cabinets(36, 180, 5),
    Cabinets(37, 187, 5),
]

internal_urls = []

teachers_list_dirty = []
teachers_list = []
our_cabinets_dirty = []
our_cabinets = []


def getTeachers(end_row, cur, col):
    files_list = os.listdir("static/zip/ОЧНИКИ")
    file_date = files_list[0].split(' ')
    #if(date_now == file_date[2]):
    excel_file = xlrd.open_workbook(f"static/zip/ОЧНИКИ/{files_list[0]}")
    #else:
    #    excel_file = xlrd.open_workbook(f"static/zip/ОЧНИКИ/{files_list[1]}")
    list1_sheet = excel_file.sheet_by_index(0)
    cursor = cur
    while cursor < end_row:
        teachers_list_dirty.append(Teachers_name(list1_sheet.cell_value(rowx=cursor, colx=col), cursor + 1, col))
        #print(list1_sheet.cell_value(rowx=cursor, colx=col), cursor + 1, col)
        cursor += 1
    
    for i in teachers_list_dirty:
        #print(f"Имя: {i.name} строка: {i.rowx} колона: {i.colx}")
        if (i.name != "" 
            and (
                i.name == "Асаева" or
                i.name == "Гончарова" or
                i.name == "Гречушкина" or
                i.name == "Арабчикова" or
                i.name == "Бакулина" or
                i.name == "Костылева" or
                i.name == "Тихонова" or
                i.name == "Кулаков" or
                i.name == "Судакова" or
                i.name == "Воробьева" 
            )):
            teachers_list.append(i)
            #print(f"Имя: {i.name} строка: {i.rowx} колона: {i.colx}")

def cabinets_mapping():
    for i in teachers_list:
        for a in cabinets:
            start_row = a.rowx_start
            end_row = start_row + 7
            while start_row < end_row:
                if (i.rowx == start_row and i.colx == a.colx):
                    our_cabinets_dirty.append([str(a.number), str(i.name)])
                start_row = start_row + 1

    for i in our_cabinets_dirty:
        if i not in our_cabinets:
            our_cabinets.append(i)

def push2Firebase():
    for i in our_cabinets:
            database.child("cabinets").push({
                    "num_cab": f"{i[0]}",
                    "lastname": f"{i[1]}"
                    })


def getData():
    bar.start()

    parsing()
    download_file(internal_urls[2])
    zipOpen()

    bar.update(1)

    getTeachers(142, 1, 2)
    cabinets_mapping()
    database.child("cabinets").remove()
    push2Firebase()

    bar.update(2)

    teachers_list_dirty.clear()
    teachers_list.clear()
    our_cabinets_dirty.clear()
    our_cabinets.clear()

    bar.update(3)

    getTeachers(142, 1, 5)
    cabinets_mapping()
    push2Firebase()

    bar.update(4)

    teachers_list_dirty.clear()
    teachers_list.clear()
    our_cabinets_dirty.clear()
    our_cabinets.clear()

    bar.update(5)

    getTeachers(193, 144, 2)
    cabinets_mapping()
    push2Firebase()

    bar.update(6)

    teachers_list_dirty.clear()
    teachers_list.clear()
    our_cabinets_dirty.clear()
    our_cabinets.clear()

    bar.update(7)

    getTeachers(193, 144, 5)
    cabinets_mapping()
    push2Firebase()

    bar.update(8)

    teachers_list_dirty.clear()
    teachers_list.clear()
    our_cabinets_dirty.clear()
    our_cabinets.clear()

    bar.finish()

timer = 10

seconds = timer
os.system('cls')

while 1==1:
    time.sleep(1)
    seconds = seconds - 1
    hours, minutes, secondsPr = seconds // 3600, (seconds - 3600 * (seconds // 3600)) // 60, seconds % 60
    sys.stdout.write("\r")
    sys.stdout.write(f"{hours:02}:{minutes:02}:{secondsPr:02} до попытки получения расписаний") 
    sys.stdout.flush()
    if(seconds == 0):
        print("\nПопытка получения расписания кабинетов")
        try:
            shutil.rmtree('/static', ignore_errors=True)
            getData()
            print("\nКабинеты обновлены")
            print("\nПопытка получения расписания уроков")
            shutil.rmtree('/static', ignore_errors=True)
            getDataTimetable()
            print("\nРасписание обновлено")
            seconds = timer
            os.system('cls')
        except:
            print("\nПроизошла ошибка обновления, производиться новая попытка")
            seconds = 1